import pytest
import os
import io
import time
import openpyxl
from fastapi.testclient import TestClient

from apps.admin_backend.main import app
from apps.admin_backend.ocr.symbol_catalog import (
    classify_tag,
    parse_line_specification
)
from apps.admin_backend.ocr.pid_graph_extractor import (
    pid_extractor,
    PIDGraphPayload
)
from apps.admin_backend.sovereignty.tamper_log import audit_log

client = TestClient(app)

# ==============================================================================
# TEST SUITE: P&ID SYMBOL DETECTION & TOPOLOGICAL GRAPH EXTRACTION (SIH PS 26117)
# ==============================================================================

def test_symbol_catalog_classification():
    """Verifies that ISA-5.1 tags and ISO 10628 equipment codes are classified accurately."""
    # 1. Major Equipment Tags
    pump = classify_tag("P-101A")
    assert pump["is_recognized"] is True
    assert pump["category"] == "PUMP"
    assert pump["tag"] == "P-101A"

    tank = classify_tag("TK-201")
    assert tank["is_recognized"] is True
    assert tank["category"] == "TANK"

    col = classify_tag("C-101")
    assert col["is_recognized"] is True
    assert col["category"] == "COLUMN"

    # 2. Valves and Actuators
    cv = classify_tag("FV-1002")
    assert cv["is_recognized"] is True
    assert cv["category"] == "VALVE"

    esdv = classify_tag("ESDV-101")
    assert esdv["is_recognized"] is True
    assert esdv["category"] == "VALVE"

    # 3. Instrumentation
    pt = classify_tag("PT-1002")
    assert pt["is_recognized"] is True
    assert pt["category"] == "INSTRUMENT"
    assert "Pressure" in pt["description"]

    # 4. Pipeline Specifications
    line = parse_line_specification('6"-HC-1001-CS-150#')
    assert line is not None
    assert line["pipe_size"] == "6 inch"
    assert line["fluid_service"] == "HC"
    assert line["material_class"] == "CS"
    assert "150#" in line["flange_rating"]

def test_pid_graph_construction_and_stats():
    """Verifies topological graph structure for canonical refinery CDU schematic."""
    graph = pid_extractor.build_synthetic_demo_pid("PID-TEST-CDU")
    assert graph is not None
    assert graph.pid_id == "PID-TEST-CDU"
    assert graph.total_equipment >= 3
    assert graph.total_valves >= 4
    assert graph.total_instruments >= 2
    assert graph.total_lines >= 6

    # Verify NetworkX representation
    G = pid_extractor.to_networkx(graph)
    assert G.number_of_nodes() == len(graph.nodes)
    assert G.number_of_edges() == len(graph.edges)

def test_pid_path_tracing_from_tank_to_column():
    """
    Traces process fluid connection from Feed Tank TK-101 to Distillation Column C-101.
    Verifies that the ordered path captures all intermediate isolation and control valves.
    """
    graph = pid_extractor.build_synthetic_demo_pid("PID-TEST-CDU")
    trace = pid_extractor.find_path(graph, "TK-101", "C-101")

    assert trace.path_exists is True
    assert trace.total_hops >= 5
    assert len(trace.steps) >= 6

    step_tags = [s["tag"] for s in trace.steps]
    assert "TK-101" in step_tags
    assert "P-101A" in step_tags
    assert "FV-1002" in step_tags
    assert "ESDV-101" in step_tags
    assert "C-101" in step_tags

    # Check that valves are captured
    assert len(trace.valves_in_path) >= 3

def test_pid_upstream_isolation_valves_for_loto():
    """
    Verifies that querying upstream isolation points for Control Valve FV-1002
    identifies upstream check valve CV-101 and suction gate valve V-101.
    """
    graph = pid_extractor.build_synthetic_demo_pid("PID-TEST-CDU")
    iso_res = pid_extractor.get_upstream_isolation_valves(graph, "FV-1002")

    assert iso_res.is_isolated is True
    assert len(iso_res.isolation_valves) >= 2
    valve_tags = [v["tag"] for v in iso_res.isolation_valves]
    assert "V-101" in valve_tags
    assert "CV-101" in valve_tags
    assert "LOTO" in iso_res.loto_recommendation

def test_pid_excel_asset_register_generation():
    """
    Verifies multi-sheet Excel generation:
    1. Equipment Register
    2. Valve Schedule
    3. Line Connectivity Matrix
    """
    graph = pid_extractor.build_synthetic_demo_pid("PID-TEST-CDU")
    xlsx_path = pid_extractor.export_asset_matrix_excel(graph)

    assert os.path.exists(xlsx_path)
    assert xlsx_path.endswith(".xlsx")

    wb = openpyxl.load_workbook(xlsx_path)
    assert "1. Equipment Register" in wb.sheetnames
    assert "2. Valve Schedule" in wb.sheetnames
    assert "3. Line Connectivity Matrix" in wb.sheetnames

    ws_eq = wb["1. Equipment Register"]
    assert ws_eq["A1"].value == "MRPL SOVEREIGN WORKBENCH - P&ID ASSET REGISTER"
    assert ws_eq.max_row >= 5

def test_pid_fastapi_rest_endpoints():
    """
    Verifies FastAPI endpoints:
    - POST /api/ocr/pid/upload
    - GET /api/ocr/pid/{id}/graph
    - POST /api/ocr/pid/{id}/trace
    - GET /api/ocr/pid/{id}/upstream-valves
    - GET /api/ocr/pid/{id}/export-excel
    """
    test_pid_id = f"PID-REST-{int(time.time())}"
    dummy_png = b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01\x08\x06\x00\x00\x00\x1f\x15c4\x00\x00\x00\nIDATx\x9cc\x00\x01\x00\x00\x05\x00\x01\r\n-\xb4\x00\x00\x00\x00IEND\xaeB`\x82"
    file_payload = ("test_pid.png", io.BytesIO(dummy_png), "image/png")

    # 1. Test Upload
    res_upload = client.post(
        "/api/ocr/pid/upload",
        data={"pid_id": test_pid_id},
        files={"file": file_payload}
    )
    assert res_upload.status_code == 200
    upload_data = res_upload.json()
    assert upload_data["success"] is True
    assert upload_data["pid_id"] == test_pid_id
    assert upload_data["total_equipment"] >= 1

    # 2. Test Get Graph
    res_graph = client.get(f"/api/ocr/pid/{test_pid_id}/graph")
    assert res_graph.status_code == 200
    graph_data = res_graph.json()
    assert graph_data["pid_id"] == test_pid_id
    assert len(graph_data["nodes"]) >= 5

    # 3. Test Trace
    res_trace = client.post(
        f"/api/ocr/pid/{test_pid_id}/trace",
        json={"source": "TK-101", "target": "C-101"}
    )
    assert res_trace.status_code == 200
    trace_data = res_trace.json()
    assert trace_data["path_exists"] is True
    assert len(trace_data["steps"]) >= 5

    # 4. Test Upstream Valves
    res_valves = client.get(f"/api/ocr/pid/{test_pid_id}/upstream-valves?asset_tag=FV-1002")
    assert res_valves.status_code == 200
    assert res_valves.json()["is_isolated"] is True

    # 5. Test Export Excel
    res_excel = client.get(f"/api/ocr/pid/{test_pid_id}/export-excel")
    assert res_excel.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in res_excel.headers["content-type"]
