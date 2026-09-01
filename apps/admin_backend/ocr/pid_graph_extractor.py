import os
import time
import math
from typing import Dict, Any, List, Optional, Tuple, Set
from pydantic import BaseModel, Field
import networkx as nx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from apps.admin_backend.ocr.symbol_catalog import (
    classify_tag,
    calculate_bbox_center,
    euclidean_distance,
    parse_line_specification
)
from apps.admin_backend.sovereignty.tamper_log import audit_log

BASE_DATA_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "..", "data"))
OUTPUT_XLSX_DIR = os.path.join(BASE_DATA_DIR, "outputs", "xlsx")

# ==============================================================================
# 1. TOPOLOGICAL GRAPH DATA MODELS
# ==============================================================================
class PIDNode(BaseModel):
    id: str
    tag: str
    type: str  # "PUMP" | "VESSEL" | "COLUMN" | "VALVE" | "INSTRUMENT" | "LINE_JUNCTION"
    category: str  # "EQUIPMENT" | "VALVE" | "INSTRUMENT" | "JUNCTION"
    description: str = ""
    bbox: List[int] = Field(default_factory=lambda: [0, 0, 50, 50])  # [x1, y1, x2, y2]
    center: Tuple[int, int] = (25, 25)
    specs: Dict[str, Any] = Field(default_factory=dict)
    status: str = "NORMAL"

class PIDEdge(BaseModel):
    source_id: str
    target_id: str
    line_tag: Optional[str] = None
    line_type: str = "PROCESS"  # "PROCESS" | "PNEUMATIC_SIGNAL" | "ELECTRIC_SIGNAL"
    flow_direction: str = "FORWARD"
    pipe_size: Optional[str] = None
    fluid_service: Optional[str] = None

class PIDGraphPayload(BaseModel):
    pid_id: str
    drawing_title: str = "MRPL Process & Instrumentation Diagram"
    drawing_number: str = "MRPL-PID-001"
    nodes: List[PIDNode] = Field(default_factory=list)
    edges: List[PIDEdge] = Field(default_factory=list)
    total_equipment: int = 0
    total_valves: int = 0
    total_instruments: int = 0
    total_lines: int = 0
    adjacency: Dict[str, List[str]] = Field(default_factory=dict)

class PathTraceResult(BaseModel):
    source: str
    target: str
    path_exists: bool
    steps: List[Dict[str, Any]] = Field(default_factory=list)
    valves_in_path: List[str] = Field(default_factory=list)
    instruments_in_path: List[str] = Field(default_factory=list)
    total_hops: int = 0
    summary_narrative: str = ""

class UpstreamIsolationResult(BaseModel):
    asset_tag: str
    isolation_valves: List[Dict[str, Any]] = Field(default_factory=list)
    is_isolated: bool = False
    loto_recommendation: str = ""

# ==============================================================================
# 2. TOPOLOGICAL GRAPH EXTRACTOR ENGINE
# ==============================================================================
class PIDTopologicalGraphExtractor:
    """
    Extracts geometric, semantic, and topological information from engineering P&IDs:
    1. Parses OCR text labels and associates them with symbol bounding boxes.
    2. Builds a NetworkX topological graph representing pipes, equipment, valves, and instruments.
    3. Performs graph path tracing, upstream isolation valve analysis, and asset matrix exports.
    """
    def __init__(self):
        os.makedirs(OUTPUT_XLSX_DIR, exist_ok=True)
        # Active graph registry in memory: pid_id -> PIDGraphPayload
        self._graph_registry: Dict[str, PIDGraphPayload] = {}

    def register_graph(self, graph_payload: PIDGraphPayload):
        self._graph_registry[graph_payload.pid_id] = graph_payload

    def get_graph(self, pid_id: str) -> Optional[PIDGraphPayload]:
        return self._graph_registry.get(pid_id)

    def to_networkx(self, graph_payload: PIDGraphPayload) -> nx.MultiDiGraph:
        """Converts graph payload to NetworkX directed multigraph."""
        G = nx.MultiDiGraph()
        for node in graph_payload.nodes:
            G.add_node(
                node.id,
                tag=node.tag,
                type=node.type,
                category=node.category,
                description=node.description,
                center=node.center,
                specs=node.specs
            )
        for edge in graph_payload.edges:
            G.add_edge(
                edge.source_id,
                edge.target_id,
                line_tag=edge.line_tag,
                line_type=edge.line_type,
                flow_direction=edge.flow_direction,
                pipe_size=edge.pipe_size,
                fluid_service=edge.fluid_service
            )
        return G

    def build_synthetic_demo_pid(self, pid_id: str = "PID-MRPL-CDU-01") -> PIDGraphPayload:
        """
        Builds a canonical, high-fidelity industrial P&ID topological model for CDU Crude Pre-Flash:
        Tank TK-101 -> Line 6"-HC-1001 -> Gate Valve V-101 -> Crude Pump P-101A -> Check Valve CV-101
        -> Flow Control Valve FV-1002 (with FT-1002, PT-1002) -> Pre-Flash Column C-101.
        """
        nodes = [
            PIDNode(
                id="node_tk101",
                tag="TK-101",
                type="TANK",
                category="EQUIPMENT",
                description="Crude Oil Feed Storage Tank (Capacity: 50,000 m³)",
                bbox=[50, 150, 180, 320],
                center=(115, 235),
                specs={"service": "Crude Oil", "design_pressure": "Atmospheric"}
            ),
            PIDNode(
                id="node_v101",
                tag="V-101",
                type="GATE_VALVE",
                category="VALVE",
                description="Manual Suction Isolation Gate Valve (6 inch, 150#)",
                bbox=[240, 220, 270, 250],
                center=(255, 235),
                specs={"size": "6 inch", "type": "Manual Gate Valve", "rating": "150#"}
            ),
            PIDNode(
                id="node_p101a",
                tag="P-101A",
                type="PUMP",
                category="EQUIPMENT",
                description="Primary Crude Charge Centrifugal Pump (Flow: 350 m³/h, Head: 120m)",
                bbox=[330, 200, 420, 270],
                center=(375, 235),
                specs={"motor_kw": 250, "flow_m3h": 350, "head_m": 120, "service": "Crude Oil"}
            ),
            PIDNode(
                id="node_cv101",
                tag="CV-101",
                type="CHECK_VALVE",
                category="VALVE",
                description="Pump Discharge Non-Return Check Valve (NRV)",
                bbox=[470, 220, 500, 250],
                center=(485, 235),
                specs={"size": "4 inch", "type": "Non-Return Check Valve"}
            ),
            PIDNode(
                id="node_pt1002",
                tag="PT-1002",
                type="PRESSURE_TRANSMITTER",
                category="INSTRUMENT",
                description="Discharge Line Pressure Transmitter (Range: 0-25 bar)",
                bbox=[540, 140, 580, 180],
                center=(560, 160),
                specs={"range": "0-25 bar", "signal": "4-20mA HART"}
            ),
            PIDNode(
                id="node_ft1002",
                tag="FT-1002",
                type="FLOW_TRANSMITTER",
                category="INSTRUMENT",
                description="Crude Charge Orifice Flow Transmitter",
                bbox=[620, 140, 660, 180],
                center=(640, 160),
                specs={"range": "0-500 m3/h", "signal": "4-20mA HART"}
            ),
            PIDNode(
                id="node_fv1002",
                tag="FV-1002",
                type="CONTROL_VALVE",
                category="VALVE",
                description="Pneumatic Diaphragm Flow Control Valve (Fail Closed - FC)",
                bbox=[690, 215, 735, 255],
                center=(712, 235),
                specs={"size": "4 inch", "actuator": "Pneumatic Diaphragm", "fail_state": "FC"}
            ),
            PIDNode(
                id="node_esdv101",
                tag="ESDV-101",
                type="ESDV",
                category="VALVE",
                description="Emergency Column Infeed Shutdown Valve (SIL-3 Actuated)",
                bbox=[790, 215, 835, 255],
                center=(812, 235),
                specs={"size": "6 inch", "type": "SIL-3 Hydraulic Actuated", "fail_state": "FC"}
            ),
            PIDNode(
                id="node_c101",
                tag="C-101",
                type="COLUMN",
                category="EQUIPMENT",
                description="Crude Atmospheric Distillation / Pre-Flash Column (Trays: 48)",
                bbox=[910, 100, 1050, 380],
                center=(980, 240),
                specs={"height_m": 42.5, "diameter_m": 4.8, "trays": 48}
            )
        ]

        edges = [
            PIDEdge(
                source_id="node_tk101",
                target_id="node_v101",
                line_tag='6"-HC-1001-CS-150#',
                line_type="PROCESS",
                flow_direction="FORWARD",
                pipe_size="6 inch",
                fluid_service="HYDROCARBON"
            ),
            PIDEdge(
                source_id="node_v101",
                target_id="node_p101a",
                line_tag='6"-HC-1001-CS-150#',
                line_type="PROCESS",
                flow_direction="FORWARD",
                pipe_size="6 inch",
                fluid_service="HYDROCARBON"
            ),
            PIDEdge(
                source_id="node_p101a",
                target_id="node_cv101",
                line_tag='4"-HC-1002-CS-300#',
                line_type="PROCESS",
                flow_direction="FORWARD",
                pipe_size="4 inch",
                fluid_service="HYDROCARBON"
            ),
            PIDEdge(
                source_id="node_cv101",
                target_id="node_fv1002",
                line_tag='4"-HC-1002-CS-300#',
                line_type="PROCESS",
                flow_direction="FORWARD",
                pipe_size="4 inch",
                fluid_service="HYDROCARBON"
            ),
            PIDEdge(
                source_id="node_ft1002",
                target_id="node_fv1002",
                line_tag="PNEUMATIC_CONTROL_LOOP_1002",
                line_type="PNEUMATIC_SIGNAL",
                flow_direction="FORWARD",
                pipe_size=None,
                fluid_service="INSTRUMENT_AIR"
            ),
            PIDEdge(
                source_id="node_fv1002",
                target_id="node_esdv101",
                line_tag='6"-HC-1003-CS-300#',
                line_type="PROCESS",
                flow_direction="FORWARD",
                pipe_size="6 inch",
                fluid_service="HYDROCARBON"
            ),
            PIDEdge(
                source_id="node_esdv101",
                target_id="node_c101",
                line_tag='6"-HC-1003-CS-300#',
                line_type="PROCESS",
                flow_direction="FORWARD",
                pipe_size="6 inch",
                fluid_service="HYDROCARBON"
            )
        ]

        adjacency = {}
        for e in edges:
            adjacency.setdefault(e.source_id, []).append(e.target_id)

        payload = PIDGraphPayload(
            pid_id=pid_id,
            drawing_title="CDU-2 Crude Feed Pre-Flash & Charge P&ID",
            drawing_number="MRPL-CDU-0012-REV4",
            nodes=nodes,
            edges=edges,
            total_equipment=len([n for n in nodes if n.category == "EQUIPMENT"]),
            total_valves=len([n for n in nodes if n.category == "VALVE"]),
            total_instruments=len([n for n in nodes if n.category == "INSTRUMENT"]),
            total_lines=len(edges),
            adjacency=adjacency
        )

        self.register_graph(payload)
        return payload

    def find_path(
        self,
        graph_payload: PIDGraphPayload,
        source_tag: str,
        target_tag: str
    ) -> PathTraceResult:
        """
        Traces the topological process stream connection between two assets,
        listing every intervening valve, transmitter, and line segment.
        """
        G = self.to_networkx(graph_payload)
        
        # Locate node IDs by tag
        tag_to_id = {node.tag.upper(): node.id for node in graph_payload.nodes}
        id_to_node = {node.id: node for node in graph_payload.nodes}

        s_id = tag_to_id.get(source_tag.upper())
        t_id = tag_to_id.get(target_tag.upper())

        if not s_id or not t_id:
            return PathTraceResult(
                source=source_tag,
                target=target_tag,
                path_exists=False,
                summary_narrative=f"Asset tag '{source_tag if not s_id else target_tag}' not found in P&ID graph."
            )

        # Build undirected graph for connectivity tracing
        U = G.to_undirected()

        try:
            path_node_ids = nx.shortest_path(U, source=s_id, target=t_id)
        except nx.NetworkXNoPath:
            return PathTraceResult(
                source=source_tag,
                target=target_tag,
                path_exists=False,
                summary_narrative=f"No connected process piping path found between '{source_tag}' and '{target_tag}'."
            )

        steps: List[Dict[str, Any]] = []
        valves_in_path: List[str] = []
        instruments_in_path: List[str] = []

        for i, nid in enumerate(path_node_ids):
            node = id_to_node[nid]
            step_info = {
                "step_index": i + 1,
                "node_id": node.id,
                "tag": node.tag,
                "type": node.type,
                "category": node.category,
                "description": node.description
            }

            if node.category == "VALVE":
                valves_in_path.append(f"{node.tag} ({node.description})")
            elif node.category == "INSTRUMENT":
                instruments_in_path.append(f"{node.tag} ({node.description})")

            steps.append(step_info)

        narrative = (
            f"Trace from **{source_tag}** to **{target_tag}** spans {len(steps)} assets over process lines. "
            f"Path passes through **{len(valves_in_path)} valve(s)** ({', '.join(valves_in_path) if valves_in_path else 'None'}) "
            f"and **{len(instruments_in_path)} instrument loop(s)** ({', '.join(instruments_in_path) if instruments_in_path else 'None'})."
        )

        return PathTraceResult(
            source=source_tag,
            target=target_tag,
            path_exists=True,
            steps=steps,
            valves_in_path=valves_in_path,
            instruments_in_path=instruments_in_path,
            total_hops=len(steps) - 1,
            summary_narrative=narrative
        )

    def get_upstream_isolation_valves(
        self,
        graph_payload: PIDGraphPayload,
        asset_tag: str
    ) -> UpstreamIsolationResult:
        """
        Traverses upstream connections to find manual/actuated isolation valves
        required for Lockout/Tagout (LOTO) and maintenance isolation.
        """
        G = self.to_networkx(graph_payload)
        tag_to_id = {node.tag.upper(): node.id for node in graph_payload.nodes}
        id_to_node = {node.id: node for node in graph_payload.nodes}

        target_id = tag_to_id.get(asset_tag.upper())
        if not target_id:
            return UpstreamIsolationResult(
                asset_tag=asset_tag,
                is_isolated=False,
                loto_recommendation=f"Asset '{asset_tag}' not found in P&ID graph."
            )

        # Reverse directed graph to traverse upstream
        R = G.reverse(copy=True)
        upstream_nodes = nx.descendants(R, target_id)

        isolation_valves = []
        for nid in upstream_nodes:
            node = id_to_node[nid]
            if node.category == "VALVE":
                isolation_valves.append({
                    "tag": node.tag,
                    "type": node.type,
                    "description": node.description,
                    "specs": node.specs
                })

        is_isolated = len(isolation_valves) > 0
        loto_rec = (
            f"To safely isolate **{asset_tag}** for maintenance or emergency shutdown, "
            f"the following **{len(isolation_valves)} upstream valve(s)** must be closed and locked out (LOTO): "
            + ", ".join([v["tag"] for v in isolation_valves]) if is_isolated
            else f"No dedicated upstream isolation valve detected for '{asset_tag}'."
        )

        return UpstreamIsolationResult(
            asset_tag=asset_tag,
            isolation_valves=isolation_valves,
            is_isolated=is_isolated,
            loto_recommendation=loto_rec
        )

    def export_asset_matrix_excel(
        self,
        graph_payload: PIDGraphPayload,
        output_path: Optional[str] = None
    ) -> str:
        """
        Generates a multi-sheet formatted MRPL Asset Register & Line Connectivity Matrix (.xlsx).
        """
        pid_id = graph_payload.pid_id
        target_path = output_path or os.path.join(
            OUTPUT_XLSX_DIR, f"MRPL_PID_Asset_Register_{pid_id.replace('-', '_')}.xlsx"
        )

        wb = openpyxl.Workbook()
        # Default sheet
        ws_eq = wb.active
        ws_eq.title = "1. Equipment Register"

        # Styles
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        title_font = Font(name="Calibri", size=14, bold=True, color="1A365D")
        bold_font = Font(name="Calibri", size=11, bold=True)
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        # ----------------------------------------------------------------------
        # Sheet 1: Equipment Register
        # ----------------------------------------------------------------------
        ws_eq.append(["MRPL SOVEREIGN WORKBENCH - P&ID ASSET REGISTER"])
        ws_eq.append([f"Drawing: {graph_payload.drawing_title} ({graph_payload.drawing_number})"])
        ws_eq.append([])
        ws_eq.append(["Asset Tag", "Category", "Equipment Type", "Description", "Grid Center (X, Y)", "Service / Specs"])

        ws_eq["A1"].font = title_font
        ws_eq["A2"].font = bold_font

        for col in range(1, 7):
            cell = ws_eq.cell(row=4, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        eq_nodes = [n for n in graph_payload.nodes if n.category == "EQUIPMENT"]
        for row_idx, node in enumerate(eq_nodes, start=5):
            ws_eq.append([
                node.tag,
                node.category,
                node.type,
                node.description,
                f"({node.center[0]}, {node.center[1]})",
                str(node.specs)
            ])
            for col in range(1, 7):
                ws_eq.cell(row=row_idx, column=col).border = thin_border

        # ----------------------------------------------------------------------
        # Sheet 2: Valve Schedule
        # ----------------------------------------------------------------------
        ws_v = wb.create_sheet(title="2. Valve Schedule")
        ws_v.append(["Valve Tag", "Valve Type", "Description", "Piping Line Spec", "Fail State / Rating"])
        for col in range(1, 6):
            cell = ws_v.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        valves = [n for n in graph_payload.nodes if n.category == "VALVE"]
        for r_idx, v in enumerate(valves, start=2):
            ws_v.append([
                v.tag,
                v.type,
                v.description,
                v.specs.get("size", "Standard"),
                v.specs.get("fail_state", "Normal")
            ])
            for col in range(1, 6):
                ws_v.cell(row=r_idx, column=col).border = thin_border

        # ----------------------------------------------------------------------
        # Sheet 3: Line Connectivity Schedule
        # ----------------------------------------------------------------------
        ws_lines = wb.create_sheet(title="3. Line Connectivity Matrix")
        ws_lines.append(["Source Asset", "Target Asset", "Line Specification Tag", "Fluid Service", "Pipe Size", "Line Type"])
        for col in range(1, 7):
            cell = ws_lines.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r_idx, edge in enumerate(graph_payload.edges, start=2):
            ws_lines.append([
                edge.source_id.replace("node_", "").upper(),
                edge.target_id.replace("node_", "").upper(),
                edge.line_tag or "UNSPECIFIED",
                edge.fluid_service or "PROCESS",
                edge.pipe_size or "-",
                edge.line_type
            ])
            for col in range(1, 7):
                ws_lines.cell(row=r_idx, column=col).border = thin_border

        # Auto-adjust column widths
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 14)

        wb.save(target_path)
        return target_path

# Global Singleton
pid_extractor = PIDTopologicalGraphExtractor()
# Preload canonical demo P&ID
pid_extractor.build_synthetic_demo_pid("PID-MRPL-CDU-01")
