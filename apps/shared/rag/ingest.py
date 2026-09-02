import os
import io
import re
import time
import hashlib
from typing import List, Dict, Any, Optional, Union
from pypdf import PdfReader

from apps.shared.rag.embeddings import get_embedder
from apps.shared.rag.session_store import session_store
from apps.shared.rag.vector_store import chroma_store, MASTER_COLLECTION, PRIMARY_COLLECTION

BASE_DATA_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "..", "data"))
CHROMA_PERSIST_DIR = os.path.join(BASE_DATA_DIR, "chroma_db")
MRPL_DOCS_DIR = os.path.join(BASE_DATA_DIR, "mrpl_documents")
ONGC_POLICIES_DIR = os.path.join(BASE_DATA_DIR, "ongc_policies")

class DocumentIngestor:
    """
    Parses and chunks real MRPL SOPs, ONGC compliance policies, and ephemeral session uploads.
    Supports PDF, DOCX, XLSX, and TXT files with local BGE-M3 embeddings.
    """
    def __init__(self, chunk_size_chars: int = 512, overlap_chars: int = 64):
        self.chunk_size = chunk_size_chars
        self.overlap = overlap_chars
        self.embedder = get_embedder()

    def parse_pdf(self, source: Union[str, bytes]) -> List[Dict[str, Any]]:
        pages_content = []
        try:
            stream = io.BytesIO(source) if isinstance(source, bytes) else open(source, "rb")
            reader = PdfReader(stream)
            for page_idx, page in enumerate(reader.pages):
                try:
                    text = page.extract_text()
                    if text and text.strip():
                        cleaned = re.sub(r'[ \t]+', ' ', text)
                        cleaned = re.sub(r'\n{3,}', '\n\n', cleaned).strip()
                        pages_content.append({"page_number": page_idx + 1, "text": cleaned})
                except Exception as e:
                    print(f"  [PDF Page Parse Warning] p{page_idx+1}: {e}")
            if isinstance(source, str):
                stream.close()
        except Exception as e:
            print(f"[PDF Read Error]: {e}")
        return pages_content

    def parse_docx(self, source: Union[str, bytes]) -> List[Dict[str, Any]]:
        try:
            from docx import Document
            stream = io.BytesIO(source) if isinstance(source, bytes) else source
            doc = Document(stream)
            full_text = "\n".join([p.text for p in doc.paragraphs if p.text.strip()])
            return [{"page_number": 1, "text": full_text}]
        except Exception as e:
            print(f"[DOCX Read Error]: {e}")
            return []

    def parse_xlsx(self, source: Union[str, bytes]) -> List[Dict[str, Any]]:
        try:
            import openpyxl
            stream = io.BytesIO(source) if isinstance(source, bytes) else open(source, "rb")
            wb = openpyxl.load_workbook(stream, data_only=True)
            pages_content = []

            for sheet_idx, sheetname in enumerate(wb.sheetnames):
                sheet = wb[sheetname]
                rows_text = []
                for row in sheet.iter_rows(values_only=True):
                    row_vals = [str(v).strip() for v in row if v is not None and str(v).strip()]
                    if row_vals:
                        rows_text.append(" | ".join(row_vals))

                if rows_text:
                    sheet_text = f"Sheet: {sheetname}\n" + "\n".join(rows_text)
                    pages_content.append({"page_number": sheet_idx + 1, "text": sheet_text})

            if isinstance(source, str):
                stream.close()
            return pages_content
        except Exception as e:
            print(f"[XLSX Read Error]: {e}")
            return []

    def parse_txt(self, source: Union[str, bytes]) -> List[Dict[str, Any]]:
        try:
            if isinstance(source, bytes):
                content = source.decode("utf-8", errors="ignore")
            else:
                with open(source, "r", encoding="utf-8", errors="ignore") as f:
                    content = f.read()
            return [{"page_number": 1, "text": content}]
        except Exception as e:
            print(f"[TXT Read Error]: {e}")
            return []

    def parse_document(self, source: Union[str, bytes], filename: str = "") -> List[Dict[str, Any]]:
        ext = os.path.splitext(filename)[1].lower() if filename else ""
        if isinstance(source, str) and not ext:
            ext = os.path.splitext(source)[1].lower()

        if ext == ".pdf":
            return self.parse_pdf(source)
        elif ext in [".docx", ".doc"]:
            return self.parse_docx(source)
        elif ext in [".xlsx", ".xls", ".csv"]:
            return self.parse_xlsx(source)
        elif ext in [".txt", ".md", ".log"]:
            return self.parse_txt(source)
        # Default fallback to plain text parser
        return self.parse_txt(source)

    def extract_clause_metadata(self, text: str) -> str:
        patterns = [
            r'(?:Clause|Section|Article|Rule|Chapter|Item|Paragraph|Para)\s*[\.:]?\s*([0-9]+(?:\.[0-9]+)*)',
            r'([0-9]+\.[0-9]+(?:\.[0-9]+)*)\s+[A-Z][a-zA-Z\s]{3,30}',
            r'([A-Z]\.[0-9]+(?:\.[0-9]+)*)'
        ]
        for pat in patterns:
            match = re.search(pat, text, re.IGNORECASE)
            if match:
                return match.group(0).strip()
        return "General Section"

    def clean_title_from_filename(self, filename: str) -> str:
        name = os.path.splitext(filename)[0]
        name = re.sub(r'^[0-9]+_', '', name)
        name = name.replace('_', ' ').replace('-', ' ').strip()
        return name

    def chunk_text(
        self,
        text: str,
        page_number: int = 1,
        default_section: str = "General"
    ) -> List[Dict[str, Any]]:
        """
        Splits raw text into sliding token-sized chunks with clause metadata.
        """
        chunks: List[Dict[str, Any]] = []
        if not text or not text.strip():
            return chunks

        paragraphs = [p.strip() for p in text.split('\n\n') if p.strip()]
        current_chunk = ""
        current_section = default_section

        for para in paragraphs:
            # Check for section or clause header in paragraph
            clause = self.extract_clause_metadata(para[:120])
            if clause != "General Section":
                current_section = clause

            if len(current_chunk) + len(para) <= self.chunk_size:
                current_chunk += ("\n\n" + para if current_chunk else para)
            else:
                if current_chunk:
                    chunks.append({
                        "text": current_chunk,
                        "page_number": page_number,
                        "section": current_section
                    })
                # If paragraph itself is longer than chunk size, slice it
                if len(para) > self.chunk_size:
                    start = 0
                    while start < len(para):
                        end = start + self.chunk_size
                        sub_slice = para[start:end]
                        chunks.append({
                            "text": sub_slice,
                            "page_number": page_number,
                            "section": current_section
                        })
                        start += (self.chunk_size - self.overlap)
                    current_chunk = ""
                else:
                    current_chunk = para

        if current_chunk:
            chunks.append({
                "text": current_chunk,
                "page_number": page_number,
                "section": current_section
            })

        return chunks

    def ingest_session_document(
        self,
        file_bytes: bytes,
        filename: str,
        session_id: str,
        user_id: str = "operator"
    ) -> Dict[str, Any]:
        """
        Fast async-compatible ingestion pipeline for session-scoped ephemeral documents.
        """
        if not file_bytes or len(file_bytes) == 0:
            raise ValueError(f"Uploaded file '{filename}' is empty.")

        sha256 = hashlib.sha256(file_bytes).hexdigest()
        pages = self.parse_document(file_bytes, filename=filename)
        if not pages:
            raise ValueError(f"No readable text or tables could be extracted from '{filename}'.")

        all_chunks: List[Dict[str, Any]] = []
        for page_data in pages:
            page_num = page_data.get("page_number", 1)
            raw_text = page_data.get("text", "")
            page_chunks = self.chunk_text(raw_text, page_number=page_num)
            all_chunks.extend(page_chunks)

        if not all_chunks:
            raise ValueError(f"Unable to generate valid text chunks for '{filename}'.")

        # Ingest directly into session vector store
        receipt = session_store.ingest_session_chunks(
            session_id=session_id,
            user_id=user_id,
            file_name=filename,
            chunks=all_chunks,
            file_size_bytes=len(file_bytes),
            sha256_hash=sha256
        )

        receipt["pages_extracted"] = len(pages)
        receipt["sha256_checksum"] = sha256
        return receipt

    def ingest_master_sop_file(
        self,
        file_path: str,
        user_id: str = "admin"
    ) -> Dict[str, Any]:
        """
        Ingests a permanent master standard into Tier 1 Global Knowledge Base.
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Master file not found at '{file_path}'")

        filename = os.path.basename(file_path)
        pages = self.parse_document(file_path, filename=filename)
        title = self.clean_title_from_filename(filename)

        all_chunks: List[Dict[str, Any]] = []
        for page_data in pages:
            page_num = page_data.get("page_number", 1)
            raw_text = page_data.get("text", "")
            page_chunks = self.chunk_text(raw_text, page_number=page_num)
            all_chunks.extend(page_chunks)

        if not all_chunks:
            return {"success": False, "detail": "No chunks extracted"}

        ids = []
        docs = []
        metadatas = []
        for idx, ch in enumerate(all_chunks):
            cid = f"master_{title.replace(' ', '_')}_{idx+1}"
            ids.append(cid)
            docs.append(ch["text"])
            metadatas.append({
                "sop_id": title,
                "title": title,
                "clause": ch.get("section", "Section 1.0"),
                "page_number": ch.get("page_number", 1),
                "source_folder": "master_standards",
                "filename": filename
            })

        embeddings = self.embedder.embed_batch(docs)
        if chroma_store.master_collection:
            chroma_store.master_collection.upsert(
                ids=ids,
                documents=docs,
                embeddings=embeddings,
                metadatas=metadatas
            )

        return {
            "success": True,
            "filename": filename,
            "title": title,
            "chunks_indexed": len(ids),
            "pages_parsed": len(pages)
        }

# Global Ingestor Singleton
document_ingestor = DocumentIngestor()
